import os
import requests
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Read the NVD API key from the environment (set NVD_API_KEY in the gitignored .env, then
# `set -a; source .env; set +a` before running). Request a key at
# https://nvd.nist.gov/developers/request-an-api-key
API_KEY = os.environ.get("NVD_API_KEY", "")
BASE_URL = "https://services.nvd.nist.gov/rest/json/cves/2.0"
REQUEST_DELAY = 0.6

HEADERS = {"apiKey": API_KEY}

HEADER_ROW = ["CVE", "CVSS", "CVSS Severity", "CWE", "CWE Name", "Description"]

def get_cve_count(keyword):
    params = {"keywordSearch": keyword, "resultsPerPage": 1}
    response = requests.get(BASE_URL, params=params, headers=HEADERS)
    if response.status_code != 200:
        print(f"[ERROR] Failed count request for '{keyword}': {response.status_code}")
        return 0
    return response.json().get("totalResults", 0)


def get_full_cves(keyword):
    params = {"keywordSearch": keyword, "resultsPerPage": 2000}
    cve_ids = []
    start_index = 0
    total_results = 1

    while start_index < total_results:
        params["startIndex"] = start_index
        response = requests.get(BASE_URL, params=params, headers=HEADERS)
        if response.status_code != 200:
            print(f"[ERROR] Failed full fetch for '{keyword}': {response.status_code}")
            break
        data = response.json()
        total_results = data.get("totalResults", 0)
        vulnerabilities = data.get("vulnerabilities", [])
        for item in vulnerabilities:
            cve_ids.append(item["cve"]["id"])
        start_index += len(vulnerabilities)
        time.sleep(REQUEST_DELAY)

    return cve_ids


def get_cve_data(cve_id):
    response = requests.get(BASE_URL, headers=HEADERS, params={"cveId": cve_id})
    if response.status_code != 200:
        print(f"  [ERROR] {cve_id}: HTTP {response.status_code}")
        return None, None, None, None, None

    try:
        vuln = response.json()["vulnerabilities"][0]["cve"]

        cvss_score = None
        cvss_severity = None
        metrics = vuln.get("metrics", {})
        if "cvssMetricV31" in metrics:
            cvss_score = metrics["cvssMetricV31"][0]["cvssData"]["baseScore"]
            cvss_severity = metrics["cvssMetricV31"][0]["cvssData"]["baseSeverity"]
        elif "cvssMetricV30" in metrics:
            cvss_score = metrics["cvssMetricV30"][0]["cvssData"]["baseScore"]
            cvss_severity = metrics["cvssMetricV30"][0]["cvssData"]["baseSeverity"]
        elif "cvssMetricV2" in metrics:
            cvss_score = metrics["cvssMetricV2"][0]["cvssData"]["baseScore"]
            cvss_severity = metrics["cvssMetricV2"][0]["baseSeverity"]

        cwe_id = None
        weaknesses = vuln.get("weaknesses", [])
        if weaknesses:
            for desc in weaknesses[0]["description"]:
                if desc["lang"] == "en":
                    cwe_id = desc["value"]
                    break

        description = None
        for desc in vuln.get("descriptions", []):
            if desc["lang"] == "en":
                description = desc["value"]
                break

        return cvss_score, cvss_severity, cwe_id, cwe_id, description  # cwe_name = cwe_id (NVD doesn't provide names)

    except Exception as e:
        print(f"  [PARSE ERROR] {cve_id}: {e}")
        return None, None, None, None, None


def style_header_row(ws):
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    col_widths = [20, 8, 14, 16, 16, 60]

    for col_idx, (header, width) in enumerate(zip(HEADER_ROW, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 20


def write_cve_row(ws, row, cve_id, cvss, severity, cwe_id, cwe_name, description):
    data_font = Font(name="Arial", size=10)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(horizontal="center", vertical="top")

    values = [cve_id, cvss, severity, cwe_id, cwe_name, description]
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = data_font
        cell.alignment = wrap_align if col_idx == 6 else center_align


if __name__ == "__main__":
    filename = input("Enter output filename (without extension): ").strip()
    if not filename:
        filename = "CVE_Output"
    output_path = f"{filename}.xlsx"

    user_input = input("Enter keywords (comma-separated): ")
    keywords = [k.strip() for k in user_input.split(",") if k.strip()]

    print("\n--- CVE Counts ---")
    counts = {}
    for keyword in keywords:
        count = get_cve_count(keyword)
        counts[keyword] = count
        print(f"  {keyword}: {count} CVEs")
        time.sleep(REQUEST_DELAY)

    keywords_to_fetch = []
    for keyword in keywords:
        if counts[keyword] == 0:
            print(f"\n[SKIP] No CVEs found for '{keyword}'")
            continue
        choice = input(f"\nFetch full CVE list for '{keyword}' ({counts[keyword]} CVEs)? (y/n): ").lower()
        if choice == "y":
            keywords_to_fetch.append(keyword)
        else:
            print(f"[-] Skipping '{keyword}'")

    if not keywords_to_fetch:
        print("\nNo keywords selected. Exiting.")
        exit()

    wb = Workbook()
    wb.remove(wb.active)  # Remove default empty sheet

    for keyword in keywords_to_fetch:
        print(f"\n[+] Fetching CVE list for '{keyword}'...")
        cve_ids = get_full_cves(keyword)
        print(f"    Found {len(cve_ids)} CVEs. Enriching with NVD details...")

        ws = wb.create_sheet(title=keyword[:31])  # Excel sheet names max 31 chars
        style_header_row(ws)

        for row_idx, cve_id in enumerate(cve_ids, start=2):
            print(f"    Querying {cve_id} ({row_idx - 1}/{len(cve_ids)})...")
            cvss, severity, cwe_id, cwe_name, description = get_cve_data(cve_id)
            write_cve_row(ws, row_idx, cve_id, cvss, severity, cwe_id, cwe_name, description)
            time.sleep(REQUEST_DELAY)

        ws.freeze_panes = "A2"
        print(f"    Sheet '{keyword}' complete.")

    wb.save(output_path)
    print(f"\nDone. File saved as: {output_path}")